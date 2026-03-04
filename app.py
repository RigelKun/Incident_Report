import os
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO
from pathlib import Path

from flask import Flask, flash, redirect, render_template, request, send_file, session, url_for
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties
from werkzeug.security import check_password_hash, generate_password_hash


app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-key-change-me")
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///disaster.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)
app.config["SESSION_COOKIE_HTTPONLY"] = True

db = SQLAlchemy(app)

ALLOWED_EMERGENCY_TYPES = {"OBGYN", "Trauma", "Medical"}
ALLOWED_USER_ROLES = {"admin", "staff"}


class Incident(db.Model):
	id = db.Column(db.Integer, primary_key=True, autoincrement=True)
	emergency_type = db.Column(db.String(120), nullable=False)
	incident_name = db.Column(db.String(200), nullable=False)
	incident_date = db.Column(db.Date, nullable=False)
	incident_time = db.Column(db.Time, nullable=False)
	place = db.Column(db.String(200), nullable=False)
	driver = db.Column(db.String(120), nullable=False)
	ptv_number = db.Column(db.String(50), nullable=False)
	responders = db.Column(db.String(250), nullable=False)
	remarks = db.Column(db.Text, nullable=True)
	created_at = db.Column(db.DateTime, default=datetime.utcnow)


class User(db.Model):
	id = db.Column(db.Integer, primary_key=True, autoincrement=True)
	username = db.Column(db.String(80), unique=True, nullable=False)
	password_hash = db.Column(db.String(255), nullable=False)
	role = db.Column(db.String(20), nullable=False, default="staff")
	created_at = db.Column(db.DateTime, default=datetime.utcnow)

	def set_password(self, raw_password):
		self.password_hash = generate_password_hash(raw_password)

	def check_password(self, raw_password):
		return check_password_hash(self.password_hash, raw_password)


def get_current_user():
	user_id = session.get("user_id")
	if not user_id:
		return None
	return User.query.get(user_id)


def _is_safe_next_url(next_url):
	return bool(next_url and next_url.startswith("/"))


def login_required(view_function):
	@wraps(view_function)
	def wrapped_view(*args, **kwargs):
		if not get_current_user():
			flash("Please log in to continue.", "error")
			return redirect(url_for("login", next=request.path))
		return view_function(*args, **kwargs)

	return wrapped_view


def admin_required(view_function):
	@wraps(view_function)
	def wrapped_view(*args, **kwargs):
		user = get_current_user()
		if not user:
			flash("Please log in to continue.", "error")
			return redirect(url_for("login", next=request.path))
		if user.role != "admin":
			flash("Admin access is required for that action.", "error")
			return redirect(url_for("database"))
		return view_function(*args, **kwargs)

	return wrapped_view


@app.context_processor
def inject_current_user():
	user = get_current_user()
	return {"current_user": user}


@app.route("/login", methods=["GET", "POST"])
def login():
	if get_current_user():
		return redirect(url_for("home"))

	next_url = request.args.get("next", "")
	if request.method == "POST":
		username = request.form.get("username", "").strip()
		password = request.form.get("password", "")
		next_url = request.form.get("next", "")

		if not username or not password:
			flash("Please enter both username and password.", "error")
			return render_template("login.html", next_url=next_url)

		user = User.query.filter_by(username=username).first()
		if not user or not user.check_password(password):
			flash("Invalid username or password.", "error")
			return render_template("login.html", next_url=next_url)

		session.clear()
		session["user_id"] = user.id
		session.permanent = True
		flash("Logged in successfully.", "success")

		if _is_safe_next_url(next_url):
			return redirect(next_url)
		return redirect(url_for("home"))

	if not _is_safe_next_url(next_url):
		next_url = ""

	return render_template("login.html", next_url=next_url)


@app.route("/logout", methods=["POST"])
@login_required
def logout():
	session.clear()
	flash("Logged out successfully.", "success")
	return redirect(url_for("login"))


@app.route("/", methods=["GET", "POST"])
@login_required
def home():
	if request.method == "POST":
		emergency_type = request.form.get("emergency_type", "").strip()
		incident_name = request.form.get("incident_name", "").strip()
		incident_date_raw = request.form.get("incident_date", "").strip()
		incident_time_raw = request.form.get("incident_time", "").strip()
		place = request.form.get("place", "").strip()
		driver = request.form.get("driver", "").strip()
		ptv_number = request.form.get("ptv_number", "").strip()
		responders = request.form.get("responders", "").strip()
		remarks = request.form.get("remarks", "").strip()

		if not all(
			[
				emergency_type,
				incident_name,
				incident_date_raw,
				incident_time_raw,
				place,
				driver,
				ptv_number,
				responders,
			]
		):
			flash("Please fill in all required fields.", "error")
			return redirect(url_for("home"))

		if emergency_type not in ALLOWED_EMERGENCY_TYPES:
			flash("Emergency type must be OBGYN, Trauma, or Medical.", "error")
			return redirect(url_for("home"))

		try:
			incident_date = datetime.strptime(incident_date_raw, "%Y-%m-%d").date()
			incident_time = datetime.strptime(incident_time_raw, "%H:%M").time()
		except ValueError:
			flash("Invalid date/time format.", "error")
			return redirect(url_for("home"))

		new_incident = Incident(
			emergency_type=emergency_type,
			incident_name=incident_name,
			incident_date=incident_date,
			incident_time=incident_time,
			place=place,
			driver=driver,
			ptv_number=ptv_number,
			responders=responders,
			remarks=remarks,
		)
		db.session.add(new_incident)
		db.session.commit()
		flash("Incident saved successfully.", "success")
		return redirect(url_for("home"))

	return render_template("app.html")


@app.route("/database", methods=["GET"])
@login_required
def database():
	incidents = Incident.query.order_by(Incident.id.desc()).all()
	return render_template("database.html", incidents=incidents)


@app.route("/incident/<int:incident_id>/edit", methods=["GET", "POST"])
@login_required
def edit_incident(incident_id):
	incident = Incident.query.get_or_404(incident_id)

	if request.method == "POST":
		emergency_type = request.form.get("emergency_type", "").strip()
		incident_name = request.form.get("incident_name", "").strip()
		incident_date_raw = request.form.get("incident_date", "").strip()
		incident_time_raw = request.form.get("incident_time", "").strip()
		place = request.form.get("place", "").strip()
		driver = request.form.get("driver", "").strip()
		ptv_number = request.form.get("ptv_number", "").strip()
		responders = request.form.get("responders", "").strip()
		remarks = request.form.get("remarks", "").strip()

		if not all(
			[
				emergency_type,
				incident_name,
				incident_date_raw,
				incident_time_raw,
				place,
				driver,
				ptv_number,
				responders,
			]
		):
			flash("Please fill in all required fields.", "error")
			return redirect(url_for("edit_incident", incident_id=incident.id))

		if emergency_type not in ALLOWED_EMERGENCY_TYPES:
			flash("Emergency type must be OBGYN, Trauma, or Medical.", "error")
			return redirect(url_for("edit_incident", incident_id=incident.id))

		try:
			incident_date = datetime.strptime(incident_date_raw, "%Y-%m-%d").date()
			incident_time = datetime.strptime(incident_time_raw, "%H:%M").time()
		except ValueError:
			flash("Invalid date/time format.", "error")
			return redirect(url_for("edit_incident", incident_id=incident.id))

		incident.emergency_type = emergency_type
		incident.incident_name = incident_name
		incident.incident_date = incident_date
		incident.incident_time = incident_time
		incident.place = place
		incident.driver = driver
		incident.ptv_number = ptv_number
		incident.responders = responders
		incident.remarks = remarks

		db.session.commit()
		flash("Incident updated successfully.", "success")
		return redirect(url_for("database"))

	return render_template("edit_incident.html", incident=incident)


@app.route("/incidents/remove-all", methods=["POST"])
@admin_required
def remove_all_incidents():
	deleted_count = db.session.query(Incident).delete()
	db.session.commit()
	flash(f"{deleted_count} incident(s) deleted.", "success")
	return redirect(url_for("database"))


@app.route("/incident/<int:incident_id>/delete", methods=["POST"])
@admin_required
def delete_incident(incident_id):
	incident = Incident.query.get_or_404(incident_id)
	db.session.delete(incident)
	db.session.commit()
	flash("Incident deleted successfully.", "success")
	return redirect(url_for("database"))


@app.route("/export")
@login_required
def export_excel():
	incidents = Incident.query.order_by(Incident.id.asc()).all()
	base_dir = Path(__file__).resolve().parent
	template_path = base_dir / "instance" / "logsheet_template.xlsx"

	if template_path.exists():
		workbook = load_workbook(template_path)
		worksheet = workbook.active
	else:
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.title = "MDRRMO Logsheet"

		worksheet.merge_cells("A1:J1")
		worksheet["A1"] = "Republic of the Philippines"
		worksheet["A1"].font = Font(bold=True, size=12)
		worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

		worksheet.merge_cells("A2:J2")
		worksheet["A2"] = "Province of Camarines Sur"
		worksheet["A2"].font = Font(size=11)
		worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

		worksheet.merge_cells("A3:J3")
		worksheet["A3"] = "Municipality of Bato"
		worksheet["A3"].font = Font(size=12)
		worksheet["A3"].alignment = Alignment(horizontal="center", vertical="center")

		worksheet.merge_cells("A4:J4")
		worksheet["A4"] = "MUNICIPAL DISASTER RISK REDUCTION AND MANAGEMENT OFFICE"
		worksheet["A4"].font = Font(bold=True, size=16)
		worksheet["A4"].alignment = Alignment(horizontal="center", vertical="center")

		worksheet.merge_cells("A5:J5")
		worksheet["A5"] = "Landline No. 773 | Mobile Nos : 09154214021/09850837314 | Email Address: mdrrmolgubato@gmail.com"
		worksheet["A5"].font = Font(bold=True, size=11)
		worksheet["A5"].alignment = Alignment(horizontal="center", vertical="center")

		worksheet.merge_cells("A6:J6")
		worksheet["A6"] = "24/7 OPERATION"
		worksheet["A6"].font = Font(bold=True, size=12)
		worksheet["A6"].alignment = Alignment(horizontal="center", vertical="center")

	thin = Side(style="thin", color="000000")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	header_fill = PatternFill(fill_type="solid", fgColor="B8C2D6")
	center = Alignment(horizontal="center", vertical="center", wrap_text=True)

	column_widths = {
		"A": 6,
		"B": 20,
		"C": 28,
		"D": 16,
		"E": 14,
		"F": 26,
		"G": 16,
		"H": 12,
		"I": 28,
		"J": 24,
	}
	for col, width in column_widths.items():
		worksheet.column_dimensions[col].width = width

	headers = [
		"NO.",
		"TYPE OF EMERGENCY",
		"NAME OF INCIDENT",
		"DATE OF INCIDENT",
		"TIME OF INCIDENT",
		"PLACE OF INCIDENT",
		"DRIVER",
		"PTV",
		"RESPONDERS",
		"REMARKS",
	]
	header_row = 9
	data_start_row = 10
	for index, header_text in enumerate(headers, start=1):
		cell = worksheet.cell(row=header_row, column=index)
		if not cell.value:
			cell.value = header_text
		if not template_path.exists():
			cell.font = Font(name="Arial", bold=True, size=10)
			cell.fill = header_fill
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
			cell.border = border
		else:
			cell.font = Font(name="Arial", bold=True, size=10)

	for row_index in range(data_start_row, max(worksheet.max_row, data_start_row) + 1):
		for col_index in range(1, 11):
			worksheet.cell(row=row_index, column=col_index, value=None)

	for row_index, incident in enumerate(incidents, start=data_start_row):
		default_table_font = Font(name="Arial", size=10)
		worksheet.cell(row=row_index, column=1, value=row_index - data_start_row + 1).alignment = center
		worksheet.cell(row=row_index, column=2, value=incident.emergency_type).alignment = center
		worksheet.cell(row=row_index, column=3, value=incident.incident_name).alignment = center
		worksheet.cell(row=row_index, column=4, value=incident.incident_date.strftime("%m-%d-%y")).alignment = center
		worksheet.cell(row=row_index, column=5, value=incident.incident_time.strftime("%I:%M %p")).alignment = center
		worksheet.cell(row=row_index, column=6, value=incident.place).alignment = center
		worksheet.cell(row=row_index, column=7, value=incident.driver).alignment = center
		worksheet.cell(row=row_index, column=8, value=incident.ptv_number).alignment = center
		worksheet.cell(row=row_index, column=9, value=incident.responders).alignment = center
		worksheet.cell(row=row_index, column=10, value=incident.remarks or "").alignment = center

		for col_index in range(1, 11):
			cell = worksheet.cell(row=row_index, column=col_index)
			cell.border = border
			cell.font = default_table_font

	if not incidents:
		empty_row = data_start_row
		for col_index in range(1, 11):
			cell = worksheet.cell(row=empty_row, column=col_index)
			cell.border = border
			cell.alignment = center
			cell.font = Font(name="Arial", size=10)

	for row_num in range(1, worksheet.max_row + 1):
		worksheet.row_dimensions[row_num].height = 22

	worksheet.row_dimensions[1].height = 30
	worksheet.row_dimensions[2].height = 30
	worksheet.row_dimensions[3].height = 30
	worksheet.row_dimensions[4].height = 36
	worksheet.row_dimensions[5].height = 28
	worksheet.row_dimensions[6].height = 30
	worksheet.row_dimensions[7].height = 18
	worksheet.row_dimensions[8].height = 18

	worksheet.row_dimensions[header_row].height = 38

	worksheet.freeze_panes = "A10"
	worksheet.print_title_rows = "1:9"
	worksheet.sheet_view.zoomScale = 90
	last_print_row = max(data_start_row, worksheet.max_row)
	worksheet.print_area = f"A1:J{last_print_row}"
	worksheet.page_setup.orientation = "landscape"
	worksheet.page_setup.scale = None
	worksheet.page_setup.fitToWidth = 1
	worksheet.page_setup.fitToHeight = 0
	worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
	worksheet.print_options.horizontalCentered = True
	worksheet.print_options.verticalCentered = False
	worksheet.page_margins.left = 0.25
	worksheet.page_margins.right = 0.25
	worksheet.page_margins.top = 0.5
	worksheet.page_margins.bottom = 0.5

	output = BytesIO()
	workbook.save(output)
	output.seek(0)

	filename = f"incidents_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
	return send_file(
		output,
		as_attachment=True,
		download_name=filename,
		mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)


with app.app_context():
	db.create_all()
	if User.query.count() == 0:
		admin_username = os.environ.get("ADMIN_USERNAME", "admin")
		admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
		admin_role = os.environ.get("ADMIN_ROLE", "admin").lower()
		if admin_role not in ALLOWED_USER_ROLES:
			admin_role = "admin"

		admin_user = User(username=admin_username, role=admin_role)
		admin_user.set_password(admin_password)
		db.session.add(admin_user)
		db.session.commit()


if __name__ == "__main__":
	app.run(debug=True)
